import requests
import io
import datetime
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

NASDAQ_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/symdir/nasdaqlisted.txt"
OTHER_LISTED_URL = "https://www.nasdaqtrader.com/dynamic/symdir/otherlisted.txt"
NYSE_LISTED_URL = "https://datahub.io/core/nyse-other-listings/r/nyse-listed.csv"

class StockTickerFetcher:

    def __init__(self, snapshot_file='stock_tickers_last.csv', log_file='stock_tickers_log.csv'):
        self.snapshot_file = snapshot_file
        self.log_file = log_file

    def download_csv(self, url, delimiter=None):
        resp = requests.get(url)
        resp.raise_for_status()
        content = resp.text
        if delimiter:
            return pd.read_csv(io.StringIO(content), delimiter=delimiter)
        return pd.read_csv(io.StringIO(content))

    def fetch_master_lists(self):
        nasdaq_df = self.download_csv(NASDAQ_LISTED_URL, delimiter='|')
        nasdaq_df = nasdaq_df[['Symbol', 'Security Name']].rename(columns={'Security Name':'Company Name'})
        nasdaq_df['Exchange'] = 'NASDAQ'

        other_df = self.download_csv(OTHER_LISTED_URL, delimiter='|')
        other_df = other_df[['ACT Symbol','Security Name','Exchange']].rename(columns={'ACT Symbol':'Symbol','Security Name':'Company Name'})

        nyse_df = self.download_csv(NYSE_LISTED_URL)[['ACT Symbol','Company Name']].rename(columns={'ACT Symbol':'Symbol'})
        nyse_df['Exchange'] = 'NYSE'

        combined_df = pd.concat([nasdaq_df, other_df, nyse_df], ignore_index=True).drop_duplicates(subset='Symbol', keep='first')
        combined_df['Fetched Date'] = datetime.datetime.now().strftime('%Y-%m-%d')
        return combined_df

    def detect_changes(self, new_df):
        if not os.path.exists(self.snapshot_file):
            new_df.to_csv(self.snapshot_file, index=False)
            return new_df, pd.DataFrame(columns=new_df.columns.tolist()+['Old Name','Old Exchange','Status']), \
                   {'New':len(new_df),'Delisted':0,'Renamed':0,'Exchange Changed':0}

        old_df = pd.read_csv(self.snapshot_file)
        old_symbols = set(old_df['Symbol'])
        new_symbols = set(new_df['Symbol'])

        added_df = new_df[new_df['Symbol'].isin(new_symbols - old_symbols)]
        removed_df = old_df[old_df['Symbol'].isin(old_symbols - new_symbols)]
        removed_df['Status'] = 'Delisted'
        removed_df['Old Name'] = removed_df['Company Name']
        removed_df['Old Exchange'] = removed_df['Exchange']

        renamed_rows, exchange_changed_rows = [], []
        for sym in old_symbols & new_symbols:
            old_row = old_df[old_df['Symbol']==sym].iloc[0]
            new_row = new_df[new_df['Symbol']==sym].iloc[0]
            rename = old_row['Company Name'] != new_row['Company Name']
            exchange_change = old_row['Exchange'] != new_row['Exchange']

            if rename:
                row = new_row.copy()
                row['Old Name'] = old_row['Company Name']
                row['Old Exchange'] = old_row['Exchange']
                row['Status'] = 'Renamed'
                renamed_rows.append(row)
            if exchange_change and not rename:
                row = new_row.copy()
                row['Old Name'] = old_row['Company Name']
                row['Old Exchange'] = old_row['Exchange']
                row['Status'] = 'Exchange Changed'
                exchange_changed_rows.append(row)

        renamed_df = pd.DataFrame(renamed_rows) if renamed_rows else pd.DataFrame(columns=new_df.columns.tolist()+['Old Name','Old Exchange','Status'])
        exchange_changed_df = pd.DataFrame(exchange_changed_rows) if exchange_changed_rows else pd.DataFrame(columns=new_df.columns.tolist()+['Old Name','Old Exchange','Status'])

        diff_df = pd.concat([added_df.assign(Status='New', **{'Old Name':'','Old Exchange':''}),
                             removed_df,
                             renamed_df,
                             exchange_changed_df], ignore_index=True, sort=False)

        counts = {
            'New': len(added_df),
            'Delisted': len(removed_df),
            'Renamed': len(renamed_df),
            'Exchange Changed': len(exchange_changed_df)
        }
        return new_df, diff_df, counts

    def save_master_and_diff(self, new_df, diff_df, counts):
        today_str = datetime.datetime.now().strftime('%Y-%m-%d')
        master_file = f'stock_tickers_{today_str}.csv'
        diff_csv_file = f'stock_tickers_diff_{today_str}.csv'
        diff_xlsx_file = f'stock_tickers_diff_{today_str}.xlsx'

        new_df.to_csv(master_file, index=False)
        diff_df.to_csv(diff_csv_file, index=False)
        new_df.to_csv(self.snapshot_file, index=False)

        if not diff_df.empty:
            diff_df.to_excel(diff_xlsx_file, index=False)
            wb = load_workbook(diff_xlsx_file)
            ws = wb.active
            fills = {
                'New': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'Delisted': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'Renamed': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'Exchange Changed': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
            }

            status_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == 'Status':
                    status_col = idx
                    break

            if status_col:
                for row in range(2, ws.max_row+1):
                    status_val = ws.cell(row=row, column=status_col).value
                    if status_val in fills:
                        for col in range(1, ws.max_column+1):
                            ws.cell(row=row, column=col).fill = fills[status_val]

            wb.save(diff_xlsx_file)

        log_entry = pd.DataFrame([{
            'Date': today_str,
            'Total Symbols': len(new_df),
            **counts
        }])
        if os.path.exists(self.log_file):
            log_entry.to_csv(self.log_file, mode='a', header=False, index=False)
        else:
            log_entry.to_csv(self.log_file, index=False)

        # JSONL export
        new_df.to_json("stock_tickers_latest.jsonl", orient="records", lines=True)

        print(f"Master list: {master_file}")
        print(f"Diff CSV: {diff_csv_file}")
        print(f"Diff XLSX: {diff_xlsx_file}")
        print(f"Counts: {counts}")
        print(f"Log updated: {self.log_file}")

    def run(self):
        master_df = self.fetch_master_lists()
        updated_master, diff_df, counts = self.detect_changes(master_df)
        self.save_master_and_diff(updated_master, diff_df, counts)

if __name__ == "__main__":
    fetcher = StockTickerFetcher()
    fetcher.run()